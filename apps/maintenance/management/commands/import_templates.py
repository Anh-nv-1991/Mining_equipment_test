import openpyxl
from django.core.management.base import BaseCommand, CommandError
from apps.equipment_management.models import EquipmentCategories
from apps.maintenance.models import (
    ReplacementTemplate,
    SupplementTemplate,
    InspectionTemplate,
    CleaningTemplate,
)
from apps.maintenance.models import MAINTENANCE_LEVEL_CHOICES

# Mapping sheet name -> (Model, các cột bắt buộc trong file Excel)
# Lưu ý: Không còn lấy cột stt vì hệ thống sẽ tự tính toán thứ tự khi gộp danh sách.
# Chúng ta cũng thêm cột "inherit" nếu có.
TEMPLATE_MAP = {
    'replacement': {
        'model': ReplacementTemplate,
        'fields': [
            'category', 'maintenance_level',
            'task_name', 'replacement_type','quantity', 'manufacturer_id', 'alternative_id',
            'inherit'
        ],
    },
    'supplement': {
        'model': SupplementTemplate,
        'fields': [
            'category', 'maintenance_level', 'change_type', 'position', 'quantity',
            'inherit'
        ],
    },
    'inspection': {
        'model': InspectionTemplate,
        'fields': ['category', 'maintenance_level', 'task_name', 'inherit'],
    },
    'cleaning': {
        'model': CleaningTemplate,
        'fields': ['category', 'maintenance_level', 'task_name', 'inherit'],
    },
}


class Command(BaseCommand):
    help = "Import maintenance templates from an Excel file with 4 sheets: Replacement, Supplement, Inspection, Cleaning"

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to the Excel (.xlsx) file')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            raise CommandError(f"Error loading Excel file: {e}")

        total_imported = 0

        for sheet_name in wb.sheetnames:
            key = sheet_name.lower().strip()
            if key not in TEMPLATE_MAP:
                self.stdout.write(self.style.WARNING(f"Skipping sheet '{sheet_name}' (not mapped)."))
                continue

            self.stdout.write(self.style.SUCCESS(f"Processing sheet '{sheet_name}'..."))
            ws = wb[sheet_name]

            # Lấy header (dòng đầu tiên), chuyển về chữ thường và loại bỏ khoảng trắng thừa.
            headers = [
                str(cell.value).strip().lower() if cell.value is not None else ''
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            expected_fields = TEMPLATE_MAP[key]['fields']
            missing = [field for field in expected_fields if field not in headers]
            if missing:
                self.stdout.write(self.style.ERROR(
                    f"Sheet '{sheet_name}' is missing columns: {', '.join(missing)}"
                ))
                continue

            Model = TEMPLATE_MAP[key]['model']
            imported_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                # Bỏ qua dòng trống hoàn toàn
                if all(value is None or str(value).strip() == '' for value in row_data.values()):
                    continue

                # --- Xử lý category ---
                category_name = str(row_data.get('category', '')).strip()
                if not category_name:
                    self.stdout.write(self.style.WARNING(f"Missing category in row: {row_data}"))
                    continue
                valid_categories = set(EquipmentCategories.objects.values_list('name', flat=True))
                if category_name not in valid_categories:
                    self.stdout.write(self.style.WARNING(
                        f"Invalid category '{category_name}' in row: {row_data}"
                    ))
                    continue
                category, _ = EquipmentCategories.objects.get_or_create(name=category_name)

                # --- Xử lý maintenance_level ---
                try:
                    level_val = int(str(row_data.get('maintenance_level')).strip())
                except (ValueError, AttributeError):
                    self.stdout.write(self.style.WARNING(f"Invalid maintenance_level in row: {row_data}"))
                    continue
                valid_levels = [lvl[0] for lvl in MAINTENANCE_LEVEL_CHOICES]
                if level_val not in valid_levels:
                    self.stdout.write(self.style.ERROR(
                        f"Invalid maintenance level {level_val} in row: {row_data}"
                    ))
                    continue

                # --- Xử lý cột inherit ---
                inherit_raw = row_data.get('inherit', '')
                # Chuyển thành boolean: True nếu giá trị nằm trong ["yes", "y"], False nếu không hoặc trống.
                inherit = str(inherit_raw).strip().lower() in ["yes", "y"]

                # Xây dựng dictionary cho defaults (cho các trường có thể cập nhật)
                defaults = {'inherit': inherit}

                # Xử lý theo từng loại sheet:
                if key == 'replacement':
                    task_name = str(row_data.get('task_name', '')).strip()
                    if not task_name:
                        self.stdout.write(self.style.WARNING(f"Missing task_name in row: {row_data}"))
                        continue
                    replacement_type = str(row_data.get('replacement_type', '')).strip()
                    quantity = int(str(row_data.get('quantity', '')).strip()) if str(
                    row_data.get('quantity', '')).isdigit() else 1
                    manufacturer_id = str(row_data.get('manufacturer_id', '')).strip()
                    alternative_id = str(row_data.get('alternative_id', '')).strip()
                    defaults.update({
                        'task_name': task_name,
                        'replacement_type': replacement_type,
                        'quantity': quantity,
                        'manufacturer_id': manufacturer_id,
                        'alternative_id': alternative_id,
                    })
                    # Khóa duy nhất: (category, maintenance_level, task_name)
                    key_fields = {'category': category, 'maintenance_level': level_val, 'task_name': task_name}
                elif key == 'supplement':
                    change_type = str(row_data.get('change_type', '')).strip()
                    if not change_type:
                        self.stdout.write(self.style.WARNING(f"Missing change_type in row: {row_data}"))
                        continue
                    position = str(row_data.get('position', '')).strip() if row_data.get('position') is not None else ""
                    quantity_str = str(row_data.get('quantity', '')).strip()
                    quantity = int(quantity_str) if quantity_str.isdigit() else 1
                    defaults.update({
                        'change_type': change_type,
                        'position': position,
                        'quantity': quantity,
                    })
                    # Khóa duy nhất cho Supplement: (category, maintenance_level, change_type, position)
                    key_fields = {'category': category, 'maintenance_level': level_val, 'change_type': change_type,
                                  'position': position}
                elif key in ['inspection', 'cleaning']:
                    task_name = str(row_data.get('task_name', '')).strip()
                    if not task_name:
                        self.stdout.write(self.style.WARNING(f"Missing task_name in row: {row_data}"))
                        continue
                    defaults['task_name'] = task_name
                    # Khóa duy nhất: (category, maintenance_level, task_name)
                    key_fields = {'category': category, 'maintenance_level': level_val, 'task_name': task_name}
                else:
                    # Nếu không thuộc sheet nào (chưa xảy ra)
                    continue

                # Cập nhật hoặc tạo mới template dựa trên khóa duy nhất đã xác định
                obj, created = Model.objects.update_or_create(
                    **key_fields,
                    defaults=defaults
                )
                imported_count += 1

            self.stdout.write(self.style.SUCCESS(
                f"Imported {imported_count} records from sheet '{sheet_name}'."
            ))
            total_imported += imported_count

        self.stdout.write(self.style.SUCCESS(
            f"Total imported records: {total_imported}"
        ))
